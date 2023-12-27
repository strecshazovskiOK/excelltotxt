from openpyxl import workbook, load_workbook
from datetime import datetime
insert = "INSERT INTO"
def player():
    workbook = load_workbook("SkirmishDataset.xlsx")
    worksheet = workbook['Player']
    playerList = []
    for row in worksheet.iter_rows(min_row=2, max_row=10, min_col=1, max_col=5):
        val = (f"\n{insert} Player (player_id, nickname, email, password, money)"
               f"\nVALUES ({row[0].value},'{row[1].value}','{row[2].value}','{row[3].value}',{row[4].value});")
        playerList.append(val)
    return playerList
def tank():
    workbook = load_workbook("SkirmishDataset.xlsx")
    worksheet = workbook['Tank']
    tankList = []
    for row in worksheet.iter_rows(min_row=2, max_row=14, min_col=1, max_col=5):
        val = (f"\n{insert} Tank (tank_id, player_id , name, attack, defense)"
               f"\nVALUES ({row[0].value},{row[1].value},'{row[2].value}',{row[3].value},{row[4].value});")
        tankList.append(val)
    return tankList
def Component():
    workbook = load_workbook("SkirmishDataset.xlsx")
    worksheet = workbook['Component']
    List = []
    for row in worksheet.iter_rows(min_row=2, max_row=23, min_col=1, max_col=5):
        val = (f"\n{insert} Component (comp_id, weight, rarity, sell_price, type)"
               f"\nVALUES ({row[0].value},{row[1].value},'{row[2].value}',{row[3].value},'{row[4].value}');")
        List.append(val)
    return List
def match():
    workbook = load_workbook("SkirmishDataset.xlsx")
    worksheet = workbook['Match']
    matchList = []
    corrected_date = ''
    for row in worksheet.iter_rows(min_row=2, max_row=15, min_col=1, max_col=4):
        corrected_date = row[1].value
        if isinstance(corrected_date, datetime):
            corrected_date = datetime.replace(corrected_date,minute=10,second=10)
            print(corrected_date)
            val = (f"\n{insert} Match (match_id, date_time, duration, winner_tank)"
                   f"\nVALUES ({row[0].value},TO_DATE('{corrected_date}','YYYY-MM-DD HH:MI:SS'),{row[2].value},{row[3].value});")
        else:
            val = (f"\n{insert} Match (match_id, date_time, duration, winner_tank)"
                   f"\nVALUES ({row[0].value},TO_DATE('{row[1].value}','MM/DD/YYYY'),{row[2].value},{row[3].value});")
        matchList.append(val)
    return matchList
def TankMovement():
    workbook = load_workbook("SkirmishDataset.xlsx")
    worksheet = workbook['TankMovement']
    List = []
    for row in worksheet.iter_rows(min_row=2, max_row=22, min_col=1, max_col=6):
        val = (f"\n{insert} TankMovement (match_id, tank_id, time, position, angle, damage)"
               f"\nVALUES ({row[0].value},{row[1].value},{row[2].value},{row[3].value},{row[4].value},{row[5].value});")
        List.append(val)
    return List
def TankJoinMatch():
    workbook = load_workbook("SkirmishDataset.xlsx")
    worksheet = workbook['TankJoinMatch']
    List = []
    for row in worksheet.iter_rows(min_row=2, max_row=29, min_col=1, max_col=3):
        val = (f"\n{insert} TankJoinMatch (match_id, tank_id, mmr_point)"
               f"\nVALUES ({row[0].value},{row[1].value},{row[2].value});")
        List.append(val)
    return List
def TankComponents():
    workbook = load_workbook("SkirmishDataset.xlsx")
    worksheet = workbook['TankComponents']
    List = []
    corrected_date = ''
    for row in worksheet.iter_rows(min_row=2, max_row=23, min_col=1, max_col=3):
        corrected_date = row[2].value
        if isinstance(corrected_date, datetime):
            corrected_date = datetime.replace(corrected_date, minute=10,second=10)
            print(corrected_date)
            val = (f"\n{insert} TankComponents (tank_id, comp_id, date_attached)"
                   f"\nVALUES ({row[0].value},{row[1].value}, TO_DATE('{corrected_date}','YYYY-MM-DD HH:MI:SS'));")
        else:
            val = (f"\n{insert} TankComponents (tank_id, comp_id, date_attached)"
                   f"\nVALUES ({row[0].value},{row[1].value}, TO_DATE('{row[2].value}','MM/DD/YYYY'));")
        List.append(val)
    return List
def PlayerFriends():
    workbook = load_workbook("SkirmishDataset.xlsx")
    worksheet = workbook['PlayerFriends']
    List = []
    corrected_date = ''
    for row in worksheet.iter_rows(min_row=2, max_row=11, min_col=1, max_col=3):
        corrected_date = row[2].value
        if isinstance(corrected_date, datetime):
            corrected_date = datetime.replace(corrected_date,minute=10,second=10)
            print(corrected_date)
            val = (f"\n{insert} PlayerFriends (player_id, friend_id, date_added)"
                   f"\nVALUES ({row[0].value},{row[1].value},TO_DATE('{corrected_date}','YYYY-MM-DD HH:MI:SS'));")
        else:
            val = (f"\n{insert} PlayerFriends (player_id, friend_id, date_added)"
                   f"\nVALUES ({row[0].value},{row[1].value},TO_DATE('{row[2].value}','MM/DD/YYYY'));")
        List.append(val)
    return List
if __name__ == "__main__":
    # match()
    # print("\n\n")
    # TankComponents()
    # print("\n\n")
    # PlayerFriends()
    f = open("insert.txt", "w")
    for i in player():
        f.write(i)
    f.write('\n\n')
    for i in tank():
        f.write(i)
    f.write('\n\n')
    for i in Component():
        f.write(i)
    f.write('\n\n')
    for i in match():
        f.write(i)
    f.write('\n\n')
    for i in TankMovement():
        f.write(i)
    f.write('\n\n')
    for i in TankJoinMatch():
        f.write(i)
    f.write('\n\n')
    for i in TankComponents():
        f.write(i)
    f.write('\n\n')
    for i in PlayerFriends():
        f.write(i)
    f.close()
